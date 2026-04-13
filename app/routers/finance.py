import os
import json
import base64
import io
import time
import httpx
import calendar
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import datetime, timezone
from typing import List, Optional, Any
from pydantic import BaseModel

# Cache tip LLM (TTL 30 min)
_tip_cache: dict = {"tip": None, "type": "info", "ts": 0}

from app.database import get_db
from app.llm import call_llm_json, call_llm, _openai_keys
from app.models.wallet import Wallet
from app.models.transaction import Transaction
from app.models.fixed_charge import FixedCharge
from app.models.provisional_expense import ProvisionalExpense
from app.models.user_settings import UserSetting
from app.schemas.wallet import WalletOut, WalletCreate
from app.schemas.transaction import TransactionCreate, TransactionOut
from app.schemas.fixed_charge import FixedChargeCreate, FixedChargeOut
from app.schemas.provisional_expense import ProvisionalExpenseCreate, ProvisionalExpenseOut
from app.schemas.category_budget import CategoryBudgetCreate, CategoryBudgetPatch, CategoryBudgetOut
from app.models.category_budget import CategoryBudget

router = APIRouter(prefix="/finance", tags=["finance"])


# ─── User Settings ────────────────────────────────────────────────
def _get_settings(db: Session) -> dict:
    rows = db.query(UserSetting).all()
    return {r.key: r.value for r in rows}

@router.get("/settings")
def get_settings(db: Session = Depends(get_db)):
    return _get_settings(db)

class SettingsPatch(BaseModel):
    monthly_salary:             Optional[int] = None
    monthly_savings_goal:       Optional[int] = None
    exceptional_savings_amount: Optional[int] = None  # 0 = désactivé
    exceptional_savings_month:  Optional[int] = None  # YYYYMM, 0 = aucun
    savings_wallet_id:          Optional[int] = None  # 0 = aucun
    current_savings_balance:    Optional[int] = None  # épargne virtuelle actuelle

@router.patch("/settings")
def patch_settings(data: SettingsPatch, db: Session = Depends(get_db)):
    updates = {k: v for k, v in data.model_dump().items() if v is not None}
    for key, value in updates.items():
        row = db.query(UserSetting).filter(UserSetting.key == key).first()
        if row:
            row.value = value
        else:
            db.add(UserSetting(key=key, value=value))
    db.commit()
    return _get_settings(db)


# ─── Résumé global (widget) ───────────────────────────────────────
@router.get("/summary")
def get_summary(db: Session = Depends(get_db)):
    wallets = db.query(Wallet).all()
    total_balance = sum(w.balance for w in wallets)

    settings = _get_settings(db)
    savings_wallet_id = settings.get("savings_wallet_id", 0)
    # Épargne virtuelle (dans le même compte) ou physique (wallet séparé)
    if savings_wallet_id:
        sw = db.query(Wallet).filter(Wallet.id == savings_wallet_id).first()
        savings_balance = sw.balance if sw else 0
    else:
        savings_balance = settings.get("current_savings_balance", 0)
    usable_balance = total_balance - savings_balance

    now = datetime.now(timezone.utc)
    start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    this_month_str = now.strftime("%Y-%m")

    income = db.query(func.sum(Transaction.amount)).filter(
        Transaction.type == "income",
        Transaction.date >= start_of_month
    ).scalar() or 0

    expenses = db.query(func.sum(Transaction.amount)).filter(
        Transaction.type == "expense",
        Transaction.date >= start_of_month
    ).scalar() or 0

    # ── Budget mensuel décomposé ──────────────────────────────────
    monthly_salary       = settings.get("monthly_salary", 2_500_000)
    monthly_savings_goal = settings.get("monthly_savings_goal", 1_000_000)
    exc_amount    = settings.get("exceptional_savings_amount", 0)
    exc_month_int = settings.get("exceptional_savings_month", 0)
    this_month_int = int(now.strftime("%Y%m"))
    effective_savings = exc_amount if (exc_month_int == this_month_int and exc_amount > 0) \
                        else monthly_savings_goal

    fixed_charges = db.query(FixedCharge).filter(FixedCharge.active == True).all()
    total_fixed = sum(c.amount for c in fixed_charges)

    prov_this_month = db.query(func.sum(ProvisionalExpense.amount)).filter(
        ProvisionalExpense.month == this_month_str
    ).scalar() or 0

    cat_budgets = db.query(CategoryBudget).all()
    total_cat_budgets = sum(
        (cb.override_amount if (cb.override_month == this_month_int and cb.override_amount) else cb.default_amount)
        for cb in cat_budgets
    )
    cat_budget_items = [
        {
            "category": cb.category,
            "amount": cb.override_amount if (cb.override_month == this_month_int and cb.override_amount) else cb.default_amount,
        }
        for cb in cat_budgets
    ]

    budget_libre = monthly_salary - effective_savings - total_fixed - prov_this_month - total_cat_budgets

    return {
        "balance": total_balance,
        "savings_balance": savings_balance,
        "usable_balance": usable_balance,
        "savings_wallet_id": savings_wallet_id,
        "income": income,
        "expenses": expenses,
        "currency": "MGA",
        "budget": {
            "salary":            monthly_salary,
            "savings":           effective_savings,
            "savings_is_exc":    exc_month_int == this_month_int and exc_amount > 0,
            "fixed_charges":     total_fixed,
            "provisionals":      int(prov_this_month),
            "category_budgets":  total_cat_budgets,
            "category_budget_items": cat_budget_items,
            "libre":             budget_libre,
        },
        "wallets": [
            {"id": w.id, "name": w.name, "type": w.type, "balance": w.balance,
             "is_savings": w.id == savings_wallet_id}
            for w in wallets
        ]
    }


# ─── Portefeuilles ────────────────────────────────────────────────
@router.get("/wallets", response_model=List[WalletOut])
def get_wallets(db: Session = Depends(get_db)):
    return db.query(Wallet).all()


@router.post("/wallets", response_model=WalletOut)
def create_wallet(data: WalletCreate, db: Session = Depends(get_db)):
    wallet = Wallet(**data.model_dump())
    db.add(wallet)
    db.commit()
    db.refresh(wallet)
    return wallet


# ─── Transactions ─────────────────────────────────────────────────
@router.get("/transactions", response_model=List[TransactionOut])
def get_transactions(
    limit: int = 50,
    wallet_id: Optional[int] = None,
    db: Session = Depends(get_db)
):
    q = db.query(Transaction).order_by(Transaction.date.desc())
    if wallet_id is not None:
        q = q.filter(Transaction.wallet_id == wallet_id)
    txs = q.limit(limit).all()
    result = []
    for tx in txs:
        out = TransactionOut.model_validate(tx)
        out.wallet_name = tx.wallet.name if tx.wallet else ""
        result.append(out)
    return result


@router.get("/transactions/export")
def export_transactions_csv(db: Session = Depends(get_db)):
    txs = db.query(Transaction).order_by(Transaction.date.desc()).limit(1000).all()
    rows = ["\uFEFFDate,Type,Catégorie,Description,Montant (Ar),Portefeuille"]
    for tx in txs:
        wallet_name = tx.wallet.name if tx.wallet else ""
        sign = tx.amount if tx.type == "income" else -tx.amount
        desc = (tx.description or "").replace(",", " ")
        date_str = tx.date.strftime("%d/%m/%Y")
        rows.append(f"{date_str},{tx.type},{tx.category},{desc},{sign},{wallet_name}")
    csv_content = "\n".join(rows)
    filename = f"transactions_{datetime.now().strftime('%Y-%m-%d')}.csv"
    return StreamingResponse(
        iter([csv_content]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/transactions/export/pdf")
def export_transactions_pdf(db: Session = Depends(get_db)):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm

    txs = db.query(Transaction).order_by(Transaction.date.desc()).limit(1000).all()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        leftMargin=15*mm, rightMargin=15*mm, topMargin=15*mm, bottomMargin=15*mm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('title', parent=styles['Heading1'],
        fontSize=14, textColor=colors.HexColor('#1a1a2e'), spaceAfter=4)
    sub_style = ParagraphStyle('sub', parent=styles['Normal'],
        fontSize=9, textColor=colors.grey, spaceAfter=10)

    elems = [
        Paragraph("Transactions financières", title_style),
        Paragraph(f"Exporté le {datetime.now().strftime('%d/%m/%Y à %H:%M')} — {len(txs)} transactions", sub_style),
        Spacer(1, 4*mm),
    ]

    header = ["Date", "Type", "Catégorie", "Description", "Montant (Ar)", "Portefeuille"]
    data_rows = [header]
    for tx in txs:
        wallet_name = tx.wallet.name if tx.wallet else ""
        sign = f"+{tx.amount:,}" if tx.type == "income" else f"-{tx.amount:,}"
        desc = (tx.description or "")[:30]
        data_rows.append([
            tx.date.strftime("%d/%m/%Y"), tx.type, tx.category, desc, sign, wallet_name
        ])

    col_widths = [22*mm, 18*mm, 28*mm, 50*mm, 30*mm, 32*mm]
    tbl = Table(data_rows, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ('BACKGROUND',  (0,0), (-1,0), colors.HexColor('#1a1a2e')),
        ('TEXTCOLOR',   (0,0), (-1,0), colors.white),
        ('FONTSIZE',    (0,0), (-1,0), 8),
        ('FONTSIZE',    (0,1), (-1,-1), 7.5),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f5f5f8')]),
        ('TEXTCOLOR',   (4,1), (4,-1), colors.HexColor('#c0392b')),
        ('GRID',        (0,0), (-1,-1), 0.3, colors.HexColor('#dddddd')),
        ('ALIGN',       (4,0), (4,-1), 'RIGHT'),
        ('TOPPADDING',  (0,0), (-1,-1), 3),
        ('BOTTOMPADDING',(0,0), (-1,-1), 3),
        ('LEFTPADDING', (0,0), (-1,-1), 4),
    ]))
    # Colorer les revenus en vert
    for i, tx in enumerate(txs, start=1):
        if tx.type == "income":
            tbl.setStyle(TableStyle([('TEXTCOLOR', (4,i), (4,i), colors.HexColor('#27ae60'))]))

    elems.append(tbl)
    doc.build(elems)
    buf.seek(0)

    filename = f"transactions_{datetime.now().strftime('%Y-%m-%d')}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


class TransactionPatch(BaseModel):
    wallet_id:   Optional[int] = None
    category:    Optional[str] = None
    description: Optional[str] = None
    amount:      Optional[int] = None
    type:        Optional[str] = None

@router.patch("/transactions/{tx_id}", response_model=TransactionOut)
def patch_transaction(tx_id: int, data: TransactionPatch, db: Session = Depends(get_db)):
    tx = db.query(Transaction).filter(Transaction.id == tx_id).first()
    if not tx:
        raise HTTPException(status_code=404, detail="Transaction introuvable")

    old_wallet = db.query(Wallet).filter(Wallet.id == tx.wallet_id).first()
    # Annuler l'effet de l'ancienne transaction sur le solde
    if old_wallet:
        if tx.type == "income":  old_wallet.balance -= tx.amount
        else:                    old_wallet.balance += tx.amount

    new_wallet_id = data.wallet_id if data.wallet_id is not None else tx.wallet_id
    new_amount    = data.amount    if data.amount    is not None else tx.amount
    new_type      = data.type      if data.type      is not None else tx.type

    new_wallet = db.query(Wallet).filter(Wallet.id == new_wallet_id).first()
    if not new_wallet:
        raise HTTPException(status_code=404, detail="Nouveau portefeuille introuvable")

    # Appliquer le nouvel effet sur le solde
    if new_type == "income": new_wallet.balance += new_amount
    else:                    new_wallet.balance -= new_amount

    tx.wallet_id = new_wallet_id
    tx.amount    = new_amount
    tx.type      = new_type
    if data.category    is not None: tx.category    = data.category
    if data.description is not None: tx.description = data.description

    db.commit()
    db.refresh(tx)
    out = TransactionOut.model_validate(tx)
    out.wallet_name = new_wallet.name
    return out


@router.post("/transactions", response_model=TransactionOut)
def add_transaction(data: TransactionCreate, db: Session = Depends(get_db)):
    wallet = db.query(Wallet).filter(Wallet.id == data.wallet_id).first()
    if not wallet:
        raise HTTPException(status_code=404, detail="Portefeuille introuvable")

    if data.type == "income":
        wallet.balance += data.amount
    elif data.type == "expense":
        if wallet.balance < data.amount:
            raise HTTPException(status_code=400, detail="Solde insuffisant")
        wallet.balance -= data.amount
    else:
        raise HTTPException(status_code=400, detail="Type invalide: income ou expense")

    tx = Transaction(
        type=data.type,
        amount=data.amount,
        category=data.category,
        description=data.description,
        wallet_id=data.wallet_id,
        date=data.date or datetime.now(timezone.utc)
    )
    db.add(tx)

    # ── Auto-épargne sur réception de salaire ──────────────────────
    if data.type == "income" and data.category == "salaire":
        s = _get_settings(db)
        now_s = datetime.now(timezone.utc)
        exc_amount    = s.get("exceptional_savings_amount", 0)
        exc_month_int = s.get("exceptional_savings_month", 0)
        this_month_int = int(now_s.strftime("%Y%m"))
        eff_savings = exc_amount if (exc_month_int == this_month_int and exc_amount > 0) \
                      else s.get("monthly_savings_goal", 1_000_000)

        savings_wid = s.get("savings_wallet_id", 0)
        if savings_wid and savings_wid != data.wallet_id:
            # Wallet épargne séparé → transfert physique
            savings_wallet = db.query(Wallet).filter(Wallet.id == savings_wid).first()
            if savings_wallet and wallet.balance >= eff_savings:
                wallet.balance -= eff_savings
                savings_wallet.balance += eff_savings
                db.add(Transaction(type="expense", amount=eff_savings, category="épargne",
                    description="[Auto] Épargne mensuelle", wallet_id=wallet.id, date=now_s))
                db.add(Transaction(type="income", amount=eff_savings, category="épargne",
                    description="[Auto] Épargne mensuelle", wallet_id=savings_wallet.id, date=now_s))
        else:
            # Épargne virtuelle dans le même compte → incrémente current_savings_balance
            cur = s.get("current_savings_balance", 0)
            row = db.query(UserSetting).filter(UserSetting.key == "current_savings_balance").first()
            if row:
                row.value = cur + eff_savings
            else:
                db.add(UserSetting(key="current_savings_balance", value=cur + eff_savings))

    db.commit()
    db.refresh(tx)

    out = TransactionOut.model_validate(tx)
    out.wallet_name = wallet.name
    return out


# ─── Stats journalières du mois en cours ─────────────────────────
@router.get("/daily-stats")
def daily_stats(db: Session = Depends(get_db)):
    import calendar
    from collections import defaultdict

    now = datetime.now(timezone.utc)
    start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    today = now.day
    days_in_month = calendar.monthrange(now.year, now.month)[1]

    # ── Soldes wallets ───────────────────────────────────────────────
    settings_now    = _get_settings(db)
    wallets_total   = sum(w.balance for w in db.query(Wallet).all())
    savings_wallet_id = settings_now.get("savings_wallet_id", 0)
    if savings_wallet_id:
        sw = db.query(Wallet).filter(Wallet.id == savings_wallet_id).first()
        savings_balance = sw.balance if sw else 0
    else:
        savings_balance = settings_now.get("current_savings_balance", 0)

    txs = db.query(Transaction).filter(
        Transaction.date >= start, Transaction.date <= now
    ).all()

    daily_inc: dict = defaultdict(int)
    daily_exp: dict = defaultdict(int)
    for tx in txs:
        k = tx.date.day
        if tx.type == "income":
            daily_inc[k] += tx.amount
        elif tx.type == "expense":
            daily_exp[k] += tx.amount

    # Solde initial du mois = solde actuel - revenus du mois + dépenses du mois
    total_inc_month = sum(daily_inc.values())
    total_exp_month = sum(daily_exp.values())
    balance_start   = wallets_total - total_inc_month + total_exp_month

    days = []
    cum_inc = cum_exp = 0
    for d in range(1, today + 1):
        inc_d = daily_inc.get(d, 0)
        exp_d = daily_exp.get(d, 0)
        cum_inc += inc_d
        cum_exp += exp_d
        days.append({
            "day":          d,
            "cum_income":   cum_inc,
            "cum_expenses": cum_exp,
            "balance":      balance_start + cum_inc - cum_exp,
            "daily_exp":    exp_d,
            "daily_inc":    inc_d,
        })

    # ── Projection ──────────────────────────────────────────────────
    settings     = _get_settings(db)
    salary_amt   = settings.get("monthly_salary", 2_500_000)
    this_month_str = now.strftime("%Y-%m")

    # Taux moyen de dépense par jour (hors jours sans données)
    active_days = sum(1 for d in range(1, today + 1) if daily_exp.get(d, 0) > 0 or daily_inc.get(d, 0) > 0)
    daily_rate  = cum_exp / max(active_days, 1)

    # ── Salaire attendu ? ──────────────────────────────────────────
    salary_received = db.query(Transaction).filter(
        Transaction.type == "income",
        Transaction.category == "salaire",
        Transaction.date >= start,
    ).first() is not None

    if not salary_received:
        # Paie vers le 10, au plus tard le 15
        salary_day = 10 if today < 10 else (15 if today < 15 else None)
    else:
        salary_day = None

    # ── Charges fixes non encore payées ce mois ──────────────────
    # Séquence temporelle : salaire d'abord → charges fixes → prévisionnelles
    charges = db.query(FixedCharge).filter(FixedCharge.active == True).all()
    fixed_by_day: dict = defaultdict(int)
    # Jour de référence pour planifier les charges en retard :
    # si salaire pas encore reçu → après le jour de paie ; sinon → demain
    overdue_target = min((salary_day + 1) if (not salary_received and salary_day) else today + 1,
                         days_in_month)
    for c in charges:
        marker = f"[Auto] {c.name}"
        already = db.query(Transaction).filter(
            Transaction.description == marker,
            Transaction.date >= start,
            Transaction.wallet_id == c.wallet_id,
        ).first()
        if not already:
            if c.day_of_month > today:
                # Charge future : si elle tombe avant le salaire, la décaler après
                sched = c.day_of_month if (salary_received or not salary_day or c.day_of_month > salary_day) \
                        else salary_day + 1
            else:
                # Charge en retard → après salaire
                sched = overdue_target
            fixed_by_day[min(sched, days_in_month)] += c.amount

    # ── Dépenses prévisionnelles : étalées APRÈS les charges fixes ──
    prov_total = int(db.query(func.sum(ProvisionalExpense.amount)).filter(
        ProvisionalExpense.month == this_month_str
    ).scalar() or 0)
    # Les prévisionnelles commencent après le jour des charges (overdue_target + 1)
    prov_start = min(overdue_target + 1, days_in_month)
    remaining_prov_days = days_in_month - prov_start + 1
    prov_daily = prov_total / max(remaining_prov_days, 1)

    # ── Construction de la projection jour par jour ────────────────
    proj_exp_cum = cum_exp
    proj_inc_cum = cum_inc
    projection = []
    for d in range(today + 1, days_in_month + 1):
        fixed_d  = fixed_by_day.get(d, 0)
        salary_d = salary_amt if (salary_day and d == salary_day) else 0
        prov_d   = prov_daily if d >= prov_start else 0

        proj_exp_d    = daily_rate + fixed_d + prov_d
        proj_exp_cum += proj_exp_d
        proj_inc_cum += salary_d

        proj_balance = wallets_total + (proj_inc_cum - cum_inc) - (proj_exp_cum - cum_exp)

        projection.append({
            "day":           d,
            "proj_expenses": round(proj_exp_cum),
            "proj_income":   round(proj_inc_cum),
            "balance":       round(proj_balance),
            "daily_exp":     round(proj_exp_d),
            "daily_inc":     round(salary_d),
        })

    future_income  = proj_inc_cum - cum_inc
    future_expense = proj_exp_cum - cum_exp
    projected_end_balance = wallets_total + future_income - future_expense

    total_pending_fixed   = sum(fixed_by_day.values())
    projected_daily_spend = round(daily_rate * (days_in_month - today))
    breakdown = {
        "current_balance": wallets_total,
        "salary_incoming": round(future_income),
        "daily_spend":     projected_daily_spend,
        "fixed_charges":   total_pending_fixed,
        "provisionals":    prov_total,
    }

    return {
        "days":                  days,
        "today":                 today,
        "days_in_month":         days_in_month,
        "daily_rate":            round(daily_rate),
        "cum_income":            cum_inc,
        "cum_expenses":          cum_exp,
        "salary_day":            salary_day,
        "salary_received":       salary_received,
        "projection":            projection,
        "projected_end_balance": round(projected_end_balance),
        "breakdown":             breakdown,
        "balance_today":         wallets_total,
        "savings_balance":       savings_balance,
    }


# ─── Budget période salariale (15→15) ────────────────────────────
@router.get("/budget-period")
def budget_period(db: Session = Depends(get_db)):
    import datetime as dt
    from collections import defaultdict

    now = datetime.now(timezone.utc)

    # Déterminer la période : dernier 15 → prochain 15
    if now.day >= 15:
        ps = now.replace(day=15, hour=0, minute=0, second=0, microsecond=0)
        nm = now.month + 1 if now.month < 12 else 1
        ny = now.year if now.month < 12 else now.year + 1
        pe = now.replace(year=ny, month=nm, day=15, hour=23, minute=59, second=59, microsecond=0)
    else:
        pm = now.month - 1 if now.month > 1 else 12
        py = now.year if now.month > 1 else now.year - 1
        ps = now.replace(year=py, month=pm, day=15, hour=0, minute=0, second=0, microsecond=0)
        pe = now.replace(day=15, hour=23, minute=59, second=59, microsecond=0)

    settings = _get_settings(db)
    salary   = settings.get("monthly_salary", 2_500_000)
    savings  = settings.get("monthly_savings_goal", 1_000_000)
    exc_amt  = settings.get("exceptional_savings_amount", 0)
    exc_mo   = settings.get("exceptional_savings_month", 0)
    eff_sav  = exc_amt if (exc_mo == int(now.strftime("%Y%m")) and exc_amt > 0) else savings
    total_fixed = sum(c.amount for c in db.query(FixedCharge).filter(FixedCharge.active == True).all())
    prov_mo  = ps.strftime("%Y-%m")
    prov_amt = db.query(func.sum(ProvisionalExpense.amount)).filter(
        ProvisionalExpense.month == prov_mo).scalar() or 0
    budget_libre = salary - eff_sav - total_fixed - prov_amt

    # Dépenses réelles par jour dans la période
    txs_period = db.query(Transaction).filter(
        Transaction.type == "expense",
        Transaction.category != "épargne",
        Transaction.date >= ps,
        Transaction.date <= now,
    ).all()
    daily_exp: dict = defaultdict(int)
    for tx in txs_period:
        daily_exp[tx.date.strftime("%Y-%m-%d")] += tx.amount

    # Construire les points jour par jour
    days = []
    cur = ps.date()
    end = pe.date()
    cumulative = 0
    while cur <= end:
        day_str = cur.strftime("%Y-%m-%d")
        is_past = cur <= now.date()
        if is_past:
            cumulative += daily_exp.get(day_str, 0)
            remaining = budget_libre - cumulative
        else:
            remaining = None
        days.append({ "date": day_str, "day": cur.day, "remaining": remaining, "is_past": is_past })
        cur += dt.timedelta(days=1)

    return {
        "period_start": ps.strftime("%Y-%m-%d"),
        "period_end":   pe.strftime("%Y-%m-%d"),
        "budget_libre": budget_libre,
        "days": days,
    }


# ─── Historique budget mensuel ───────────────────────────────────
@router.get("/budget-history")
def budget_history(db: Session = Depends(get_db)):
    settings  = _get_settings(db)
    salary    = settings.get("monthly_salary", 2_500_000)
    savings   = settings.get("monthly_savings_goal", 1_000_000)
    total_fixed = sum(c.amount for c in db.query(FixedCharge).filter(FixedCharge.active == True).all())

    now = datetime.now(timezone.utc)
    months = []
    for i in range(5, -1, -1):
        y, m = now.year, now.month - i
        while m <= 0: m += 12; y -= 1
        months.append((y, m))

    result = []
    for y, m in months:
        month_str   = f"{y}-{m:02d}"
        month_start = datetime(y, m, 1, tzinfo=timezone.utc)
        last_day    = calendar.monthrange(y, m)[1]
        month_end   = datetime(y, m, last_day, 23, 59, 59, tzinfo=timezone.utc)

        depenses = db.query(func.sum(Transaction.amount)).filter(
            Transaction.type == "expense",
            Transaction.category != "épargne",
            Transaction.date >= month_start,
            Transaction.date <= month_end,
        ).scalar() or 0

        prov = db.query(func.sum(ProvisionalExpense.amount)).filter(
            ProvisionalExpense.month == month_str
        ).scalar() or 0

        labels_fr = ["jan","fév","mar","avr","mai","jun","jul","aoû","sep","oct","nov","déc"]
        result.append({
            "month": month_str,
            "label": labels_fr[m - 1],
            "budget_prevu":   salary - savings - total_fixed - prov,
            "depenses_reelles": depenses,
            "libre_reel":     salary - savings - total_fixed - depenses,
        })
    return result


# ─── Transfert entre wallets ─────────────────────────────────────
class TransferRequest(BaseModel):
    from_wallet_id: int
    to_wallet_id:   int
    amount:         int
    fee:            Optional[int] = 0
    description:    Optional[str] = None

@router.post("/transfer")
def transfer_between_wallets(data: TransferRequest, db: Session = Depends(get_db)):
    src = db.query(Wallet).filter(Wallet.id == data.from_wallet_id).first()
    dst = db.query(Wallet).filter(Wallet.id == data.to_wallet_id).first()
    if not src:
        raise HTTPException(status_code=404, detail="Wallet source introuvable")
    if not dst:
        raise HTTPException(status_code=404, detail="Wallet destination introuvable")
    fee = data.fee or 0
    total_deducted = data.amount + fee
    if src.balance < total_deducted:
        raise HTTPException(status_code=400, detail="Solde insuffisant (montant + frais)")

    desc = data.description or f"Transfert → {dst.name}"
    src.balance -= total_deducted
    dst.balance += data.amount

    now = datetime.now(timezone.utc)
    db.add(Transaction(type="expense", amount=data.amount, category="transfert",
        description=desc, wallet_id=src.id, date=now))
    db.add(Transaction(type="income", amount=data.amount, category="transfert",
        description=desc, wallet_id=dst.id, date=now))
    if fee > 0:
        db.add(Transaction(type="expense", amount=fee, category="transfert",
            description=f"Frais transfert → {dst.name}", wallet_id=src.id, date=now))
    db.commit()
    return {"ok": True, "from": src.name, "to": dst.name, "amount": data.amount, "fee": fee}


# ─── Charges fixes ────────────────────────────────────────────────
@router.get("/fixed-charges", response_model=List[FixedChargeOut])
def get_fixed_charges(db: Session = Depends(get_db)):
    charges = db.query(FixedCharge).filter(FixedCharge.active == True).all()
    result = []
    for c in charges:
        out = FixedChargeOut.model_validate(c)
        out.wallet_name = c.wallet.name if c.wallet else ""
        result.append(out)
    return result


@router.post("/fixed-charges", response_model=FixedChargeOut)
def create_fixed_charge(data: FixedChargeCreate, db: Session = Depends(get_db)):
    wallet = db.query(Wallet).filter(Wallet.id == data.wallet_id).first()
    if not wallet:
        raise HTTPException(status_code=404, detail="Portefeuille introuvable")
    charge = FixedCharge(**data.model_dump())
    db.add(charge)
    db.commit()
    db.refresh(charge)
    out = FixedChargeOut.model_validate(charge)
    out.wallet_name = wallet.name
    return out


@router.delete("/fixed-charges/{charge_id}")
def delete_fixed_charge(charge_id: int, db: Session = Depends(get_db)):
    charge = db.query(FixedCharge).filter(FixedCharge.id == charge_id).first()
    if not charge:
        raise HTTPException(status_code=404, detail="Charge introuvable")
    charge.active = False
    db.commit()
    return {"ok": True}


# ─── Dépenses prévisionnelles ─────────────────────────────────────
@router.get("/provisional-expenses", response_model=List[ProvisionalExpenseOut])
def get_provisional_expenses(month: Optional[str] = None, db: Session = Depends(get_db)):
    q = db.query(ProvisionalExpense)
    if month:
        q = q.filter(ProvisionalExpense.month == month)
    items = q.order_by(ProvisionalExpense.created_at.desc()).all()
    result = []
    for item in items:
        out = ProvisionalExpenseOut.model_validate(item)
        out.wallet_name = item.wallet.name if item.wallet else ""
        result.append(out)
    return result


@router.post("/provisional-expenses", response_model=ProvisionalExpenseOut)
def create_provisional_expense(data: ProvisionalExpenseCreate, db: Session = Depends(get_db)):
    wallet = db.query(Wallet).filter(Wallet.id == data.wallet_id).first()
    if not wallet:
        raise HTTPException(status_code=404, detail="Portefeuille introuvable")
    item = ProvisionalExpense(**data.model_dump())
    db.add(item)
    db.commit()
    db.refresh(item)
    out = ProvisionalExpenseOut.model_validate(item)
    out.wallet_name = wallet.name
    return out


@router.patch("/provisional-expenses/{item_id}")
def update_provisional_expense(item_id: int, updates: dict, db: Session = Depends(get_db)):
    item = db.query(ProvisionalExpense).filter(ProvisionalExpense.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Entrée introuvable")
    if updates.get("description"): item.description = updates["description"]
    if updates.get("amount"):      item.amount = updates["amount"]
    if updates.get("category"):    item.category = updates["category"]
    if updates.get("month"):       item.month = updates["month"]
    if updates.get("wallet_id"):   item.wallet_id = updates["wallet_id"]
    db.commit()
    db.refresh(item)
    return {"ok": True}



# ─── Paiements en attente (pour DI post-salaire) ──────────────────
@router.get("/pending-payments")
def pending_payments(db: Session = Depends(get_db)):
    now = datetime.now(timezone.utc)
    start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    month_str = now.strftime("%Y-%m")
    charges = db.query(FixedCharge).filter(FixedCharge.active == True).all()
    unpaid = []
    for c in charges:
        paid = db.query(Transaction).filter(
            Transaction.description == f"[Auto] {c.name}",
            Transaction.date >= start,
            Transaction.wallet_id == c.wallet_id,
        ).first()
        if not paid:
            unpaid.append({"id": c.id, "name": c.name, "amount": c.amount, "kind": "charge"})
    for p in db.query(ProvisionalExpense).filter(ProvisionalExpense.month == month_str).all():
        unpaid.append({"id": p.id, "name": p.description, "amount": p.amount, "kind": "provisional"})
    return {"items": unpaid}


@router.delete("/provisional-expenses/{item_id}")
def delete_provisional_expense(item_id: int, db: Session = Depends(get_db)):
    item = db.query(ProvisionalExpense).filter(ProvisionalExpense.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Entrée introuvable")
    db.delete(item)
    db.commit()
    return {"ok": True}


# ─── Budgets par catégorie ────────────────────────────────────────

@router.get("/category-budgets", response_model=List[CategoryBudgetOut])
def get_category_budgets(db: Session = Depends(get_db)):
    budgets = db.query(CategoryBudget).order_by(CategoryBudget.category).all()
    now = datetime.now(timezone.utc)
    start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    current_yyyymm = int(now.strftime("%Y%m"))
    result = []
    for b in budgets:
        effective = (
            b.override_amount
            if b.override_month == current_yyyymm and b.override_amount
            else b.default_amount
        )
        spent = db.query(func.sum(Transaction.amount)).filter(
            Transaction.type == "expense",
            Transaction.category == b.category,
            Transaction.date >= start_of_month,
        ).scalar() or 0
        result.append(CategoryBudgetOut(
            category=b.category,
            default_amount=b.default_amount,
            override_amount=b.override_amount,
            override_month=b.override_month,
            effective_limit=effective,
            spent=spent,
            ratio=round(spent / effective, 3) if effective > 0 else 0.0,
        ))
    return result


@router.post("/category-budgets", response_model=CategoryBudgetOut)
def create_category_budget(data: CategoryBudgetCreate, db: Session = Depends(get_db)):
    existing = db.query(CategoryBudget).filter(CategoryBudget.category == data.category).first()
    if existing:
        raise HTTPException(status_code=400, detail="Budget déjà défini pour cette catégorie")
    b = CategoryBudget(category=data.category, default_amount=data.default_amount)
    db.add(b)
    db.commit()
    db.refresh(b)
    return CategoryBudgetOut(
        category=b.category, default_amount=b.default_amount,
        override_amount=None, override_month=None,
        effective_limit=b.default_amount, spent=0, ratio=0.0,
    )


@router.patch("/category-budgets/{category}", response_model=CategoryBudgetOut)
def patch_category_budget(category: str, data: CategoryBudgetPatch, db: Session = Depends(get_db)):
    b = db.query(CategoryBudget).filter(CategoryBudget.category == category).first()
    if not b:
        raise HTTPException(status_code=404, detail="Budget introuvable")
    if data.default_amount is not None:
        b.default_amount = data.default_amount
    if data.override_amount is not None:
        if data.override_amount > 0:
            b.override_amount = data.override_amount
            b.override_month  = data.override_month
        else:
            b.override_amount = None
            b.override_month  = None
    db.commit()
    db.refresh(b)
    now = datetime.now(timezone.utc)
    current_yyyymm = int(now.strftime("%Y%m"))
    start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    effective = (
        b.override_amount
        if b.override_month == current_yyyymm and b.override_amount
        else b.default_amount
    )
    spent = db.query(func.sum(Transaction.amount)).filter(
        Transaction.type == "expense",
        Transaction.category == b.category,
        Transaction.date >= start_of_month,
    ).scalar() or 0
    return CategoryBudgetOut(
        category=b.category, default_amount=b.default_amount,
        override_amount=b.override_amount, override_month=b.override_month,
        effective_limit=effective, spent=spent,
        ratio=round(spent / effective, 3) if effective > 0 else 0.0,
    )


@router.delete("/category-budgets/{category}")
def delete_category_budget(category: str, db: Session = Depends(get_db)):
    b = db.query(CategoryBudget).filter(CategoryBudget.category == category).first()
    if not b:
        raise HTTPException(status_code=404, detail="Budget introuvable")
    db.delete(b)
    db.commit()
    return {"ok": True}


# ─── Auto-apply charges fixes ─────────────────────────────────────
@router.get("/apply-fixed-charges")
def apply_fixed_charges(db: Session = Depends(get_db)):
    """Applique les charges fixes du mois si on est le 15 ou après et qu'elles ne l'ont pas encore été."""
    now = datetime.now(timezone.utc)
    if now.day < 15:
        return {"applied": 0, "charges": [], "skipped": [], "reason": "Avant le 15 du mois"}

    start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    charges = db.query(FixedCharge).filter(FixedCharge.active == True).all()

    applied_names: List[str] = []
    skipped_names: List[str] = []

    for charge in charges:
        marker = f"[Auto] {charge.name}"
        already = db.query(Transaction).filter(
            Transaction.description == marker,
            Transaction.date >= start_of_month,
            Transaction.wallet_id == charge.wallet_id,
        ).first()

        if already:
            skipped_names.append(charge.name)
            continue

        wallet = db.query(Wallet).filter(Wallet.id == charge.wallet_id).first()
        if not wallet:
            skipped_names.append(charge.name)
            continue

        # Déduire du wallet
        wallet.balance -= charge.amount

        tx = Transaction(
            type="expense",
            amount=charge.amount,
            category="charge_fixe",
            description=marker,
            wallet_id=charge.wallet_id,
            date=now,
        )
        db.add(tx)
        applied_names.append(charge.name)

    db.commit()
    return {
        "applied": len(applied_names),
        "charges": applied_names,
        "skipped": skipped_names,
    }


# ─── Chat LLM ─────────────────────────────────────────────────────
class ChatRequest(BaseModel):
    message: str
    history: List[dict] = []


@router.post("/chat")
def chat(data: ChatRequest, db: Session = Depends(get_db)):
    """Interprète un message en langage naturel et exécute ou prépare une action financière."""
    wallets = db.query(Wallet).all()
    wallet_list = "\n".join(
        f'- id={w.id}, nom="{w.name}", type={w.type}, solde={w.balance} Ar'
        for w in wallets
    )

    # Dernières transactions pour le contexte (limité à 8 pour réduire la latence)
    recent_txs = (
        db.query(Transaction)
        .order_by(Transaction.date.desc())
        .limit(8)
        .all()
    )
    tx_lines = "\n".join(
        f'- [{tx.date.strftime("%d/%m/%Y")}] {tx.type} {tx.amount} Ar — {tx.category}'
        f'{(" · " + tx.description) if tx.description else ""}'
        f' ({tx.wallet.name if tx.wallet else "?"})'
        for tx in recent_txs
    )
    tx_context = f"\nDernières transactions :\n{tx_lines}" if tx_lines else ""

    # Charges fixes actives
    fixed_charges = db.query(FixedCharge).filter(FixedCharge.active == True).all()
    total_fixed = sum(c.amount for c in fixed_charges)
    fixed_lines = "\n".join(
        f'- {c.name} : {c.amount:,} Ar/mois (le {c.day_of_month}, {c.category}, wallet_id={c.wallet_id})'
        for c in fixed_charges
    )
    fixed_context = f"\nCharges fixes mensuelles (total : {total_fixed:,} Ar/mois) :\n{fixed_lines}" if fixed_charges else "\nCharges fixes : aucune"

    # Dépenses prévisionnelles en cours
    now_ctx = datetime.now(timezone.utc)
    this_month_str = now_ctx.strftime("%Y-%m")
    prov_txs = (
        db.query(ProvisionalExpense)
        .filter(ProvisionalExpense.month >= this_month_str)
        .order_by(ProvisionalExpense.month.asc())
        .limit(10)
        .all()
    )
    prov_this_month = sum(p.amount for p in prov_txs if p.month == this_month_str)
    prov_lines = "\n".join(
        f'- id={p.id} [{p.month}] {p.description} — {p.amount:,} Ar ({p.category})'
        f'{(" · wallet_id=" + str(p.wallet_id)) if p.wallet_id else ""}'
        for p in prov_txs
    )
    prov_context = f"\nDépenses prévisionnelles à venir :\n{prov_lines}" if prov_lines else "\nDépenses prévisionnelles : aucune"

    # Budgets par catégorie (pour matching alimentation, etc.)
    cat_budgets = db.query(CategoryBudget).all()
    now_yyyymm = int(now_ctx.strftime("%Y%m"))
    cat_budget_lines = "\n".join(
        f'- {b.category} : plafond {(b.override_amount if b.override_month == now_yyyymm and b.override_amount else b.default_amount):,} Ar/mois'
        for b in cat_budgets
    )
    cat_context = f"\nBudgets catégorie (dépenses suivies automatiquement par catégorie) :\n{cat_budget_lines}" if cat_budget_lines else ""

    # Profil utilisateur + budget réel
    settings = _get_settings(db)
    monthly_salary       = settings.get("monthly_salary", 2_500_000)
    monthly_savings_goal = settings.get("monthly_savings_goal", 1_000_000)
    savings_wid          = settings.get("savings_wallet_id", 0)
    savings_w            = db.query(Wallet).filter(Wallet.id == savings_wid).first() if savings_wid else None
    savings_wallet_line  = f"\nWallet épargne désigné : id={savings_w.id}, nom=\"{savings_w.name}\", solde actuel={savings_w.balance:,} Ar" if savings_w else "\nWallet épargne : non configuré"
    fmt_n = lambda n: f"{n:,}".replace(",", " ")

    # Épargne effective ce mois (exceptionnelle si définie pour ce mois, sinon défaut)
    exc_amount    = settings.get("exceptional_savings_amount", 0)
    exc_month_int = settings.get("exceptional_savings_month", 0)
    this_month_int = int(now_ctx.strftime("%Y%m"))
    exceptional_active = (exc_month_int == this_month_int and exc_amount > 0)
    effective_savings  = exc_amount if exceptional_active else monthly_savings_goal

    # Budget réel = salaire - épargne - charges fixes - prévisionnelles ce mois
    budget_apres_epargne  = monthly_salary - effective_savings
    budget_apres_charges  = budget_apres_epargne - total_fixed
    budget_libre_ce_mois  = budget_apres_charges - prov_this_month

    exc_note = (
        f"\n⚡ ÉPARGNE EXCEPTIONNELLE CE MOIS : {fmt_n(exc_amount)} Ar (au lieu de {fmt_n(monthly_savings_goal)} Ar)"
        if exceptional_active else ""
    )

    system_prompt = f"""Tu es Zoky kiontabla, gestionnaire de compte personnel expert. Tu parles à un utilisateur malgache qui gère ses finances en Ariary (Ar).

Ton style : direct, chiffré, sans langue de bois. Comme un CFO ami qui dit la vérité.
Tes réponses doivent avoir de la personnalité — pas de réponses plates et génériques.
Utilise \\n\\n entre les paragraphes pour aérer. Donne des chiffres précis, des ratios, des recommandations fermes.
Pour les questions de budget/achat : calcule immédiatement (épargne nécessaire, délai, ratio dépenses/revenus).
Pour les conseils : sois spécifique, pas "économise davantage" mais "réduis alimentation de 20% = +X Ar/mois".

RÈGLES DE WALLET — PRIORITÉ ABSOLUE (appliquer AVANT toute autre inférence) :
1. WALLET PAR DÉFAUT : si l'utilisateur ne précise pas de wallet, toujours utiliser "Argent liquide".
   Exemples : "j'ai dépensé 5000", "achat de riz 3000", "taxi 2000" → wallet = Argent liquide.
2. ALIMENTATION / NOURRITURE : TOUTES les dépenses de nourriture (repas, restaurant, courses, épicerie, riz, marché, déjeuner, dîner, petit-déjeuner, cuisine…) → wallet = "Argent liquide" OBLIGATOIREMENT.
   JAMAIS depuis Compte bancaire, Mvola ou Orange Money pour l'alimentation, SAUF si l'utilisateur précise explicitement un autre wallet.
3. CRÉDIT YAS / OPÉRATEUR YAS : "crédit Yas", "recharge Yas", "forfait Yas" → wallet = "Mvola" par défaut.
   SAUF si l'utilisateur précise une autre origine (ex: "en liquide", "cash", "avec mon compte") → utiliser le wallet précisé.
4. CONNEXION ORANGE / INTERNET ORANGE : "paiement connexion Orange", "internet Orange", "forfait Orange", "pass Internet Orange" → wallet = "Orange Money".
5. TRANSFERTS CASH / MOBILE MONEY → LIQUIDE : "j'ai retiré de l'Orange Money", "retrait Mvola", "transfert banque en liquide"
   → utiliser l'action "transfer" avec to_wallet = "Argent liquide".
6. FRAIS MVOLA : les frais de retrait ou de service Mvola → expense depuis "Mvola", catégorie "mvola".

PROFIL FINANCIER (RÈGLES ABSOLUES — NE JAMAIS DÉROGER) :
Salaire : {fmt_n(monthly_salary)} Ar/mois
Épargne mensuelle (défaut) : {fmt_n(monthly_savings_goal)} Ar/mois (NON NÉGOCIABLE){exc_note}
Décomposition ce mois :
  - Épargne ce mois (NON NÉGOCIABLE)   : -{fmt_n(effective_savings)} Ar
  - Charges fixes mensuelles           : -{fmt_n(total_fixed)} Ar
  - Dépenses prévisionnelles ce mois   : -{fmt_n(prov_this_month)} Ar
  ────────────────────────────────────────
  BUDGET RÉELLEMENT DISPONIBLE         :  {fmt_n(budget_libre_ce_mois)} Ar

⚠ RÈGLE CRITIQUE : Le budget disponible est {fmt_n(budget_libre_ce_mois)} Ar, PAS {fmt_n(budget_apres_epargne)} Ar.
  Toujours raisonner sur {fmt_n(budget_libre_ce_mois)} Ar comme plafond de dépenses du mois.

DONNÉES FINANCIÈRES ACTUELLES :{savings_wallet_line}
{wallet_list}
{fixed_context}
{prov_context}
{cat_context}
{tx_context}

Réponds UNIQUEMENT avec un JSON valide (sans markdown) :
{{
  "action": "add_transaction" | "patch_transaction" | "add_multiple_transactions" | "transfer" | "add_fixed_charge" | "add_provisional_expense" | "update_provisional_expense" | "delete_provisional_expense" | "create_wallet" | "update_settings" | "generate_devis" | "generate_invoice" | "generate_excel" | "generate_report" | "answer",
  "message": "Réponse avec \\n\\n entre paragraphes",
  "requires_confirmation": true | false,
  "data": {{ ... }} | null
}}

RÈGLES D'ACTION — CHOISIS LA BONNE :

▶ transfer → virement entre deux wallets (épargne, transfert, remboursement interne)
  Ex: "vire 1M vers mon épargne", "j'ai épargné 1.000.000 ce mois", "transfert de 500k de MVola vers Banque", "j'ai transféré 20k de Mvola vers banque avec 500 de frais"
  data = {{ "from_wallet_id": int, "to_wallet_id": int, "amount": int, "fee": int (0 si non précisé), "description": str }}
  requires_confirmation = true si wallet source ou montant incertains, sinon false

▶ patch_transaction → corriger une transaction existante (mauvais wallet, mauvais montant, mauvaise catégorie)
  Ex: "corrige la dernière dépense", "change le wallet de cette transaction", "c'était sur Mvola pas sur cash"
  data = {{ "id": int, "wallet_id": int (optionnel), "amount": int (optionnel), "category": str (optionnel), "description": str (optionnel), "type": str (optionnel) }}
  requires_confirmation = true
  IMPORTANT : utilise l'id de la transaction visible dans les dernières transactions ci-dessus

▶ add_transaction → dépense ou revenu DÉJÀ effectué (passé ou présent immédiat)
  Ex: "j'ai dépensé 5000", "j'ai reçu ma paie", "achat de ce matin"
  data = {{ "type": "income"|"expense", "amount": int, "category": str, "wallet_id": int, "description": str,
            "provisional_id": int|null, "fixed_charge_id": int|null }}
  Catégories : alimentation, transport, salaire, transfert, loyer, santé, loisirs, mvola, orange_money, internet, autre
  requires_confirmation = false si montant ET portefeuille clairs

  RÈGLES DE MATCHING AUTOMATIQUE (PRIORITÉ HAUTE) :
  • Si la dépense correspond à une dépense prévisionnelle connue (même description, même personne, même objet)
    → mettre provisional_id = <id de la provision> + utiliser wallet_id et category de la provision
    → cela marquera la provision comme réalisée (elle sera supprimée)
    Ex: provision "Mika 50k" → utilisateur dit "j'ai donné l'argent à Mika" → provisional_id = <id>

  • Si la dépense correspond à une charge fixe connue (loyer, connexion Orange, Yas, lessive…)
    → mettre fixed_charge_id = <id de la charge> + utiliser son wallet_id
    → cela marquera la charge comme payée ce mois
    Ex: "j'ai payé le loyer" → fixed_charge_id = <id du loyer>

  • Si la catégorie correspond à un budget catégorie (alimentation, transport…)
    → simplement mettre la bonne category (le suivi du budget est automatique)
    Ex: "achat de riz 30k" → category = "alimentation" (déduit du budget alimentation)

▶ add_multiple_transactions → plusieurs transactions passées d'un coup
  data = {{ "transactions": [...] }}

▶ add_provisional_expense → dépense FUTURE prévue, ponctuelle (pas récurrente)
  Ex: "je vais devoir payer X", "je dois donner Y à Z la semaine prochaine", "prévoir X pour le mois prochain"
  UTILISE CECI si quelqu'un dit "après la prochaine paie", "prochainement", "prévoir", "je vais donner"
  data = {{ "description": str, "amount": int, "wallet_id": int, "category": str, "month": "YYYY-MM" }}
  month = mois prévu (mois prochain si non précisé)
  requires_confirmation = true

▶ update_provisional_expense → modifier une dépense prévisionnelle existante
  Ex: "change X par Y", "c'est 110k pas 11k", "modifie la description de..."
  Utilise l'id visible dans les dépenses prévisionnelles ci-dessus
  data = {{ "id": int, "description": str (optionnel), "amount": int (optionnel), "category": str (optionnel), "month": str (optionnel) }}
  requires_confirmation = true

▶ delete_provisional_expense → supprimer une dépense prévisionnelle
  data = {{ "id": int }}
  requires_confirmation = true

▶ add_fixed_charge → charge qui REVIENT CHAQUE MOIS (loyer, abonnement, internet, etc.)
  Ex: "loyer de 300k par mois", "abonnement Canal+", "prélèvement mensuel"
  ATTENTION : n'utilise PAS ceci pour une dépense ponctuelle même future
  data = {{ "name": str, "amount": int, "wallet_id": int, "category": str, "day_of_month": int }}
  requires_confirmation = true

▶ update_settings → modifier salaire, objectif épargne, ou épargne exceptionnelle ce mois
  Ex: "ce mois j'épargne 1.2M", "mon salaire est maintenant 3M", "augmente mon objectif d'épargne à 1.2M"
  data = {{
    "monthly_salary": int (optionnel),
    "monthly_savings_goal": int (optionnel),
    "exceptional_savings_amount": int (optionnel — montant exceptionnel CE mois uniquement),
    "exceptional_savings_month": int (optionnel — YYYYMM, ex: {this_month_int}, toujours fournir si exceptional_savings_amount fourni)
  }}
  requires_confirmation = true

▶ create_wallet : data = {{ "name": str, "type": "bank"|"mobile_money"|"cash", "balance": int }}

▶ generate_invoice / generate_devis : data = {{ "title": str, "client": str, "items": [{{"name": str, "qty": float, "unit_price": int}}], "notes": str, "type": "facture"|"devis" }}

▶ generate_excel : Export tableur. data = {{ "type": "transactions"|"wallets"|"full", "months": int }}
▶ generate_report : Rapport PDF. data = {{ "type": "transactions"|"monthly"|"full", "months": int, "title": str }}

▶ answer : Questions, analyses, conseils. Expert, chiffré, direct.

⚠ RAPPEL FORMAT — ABSOLUMENT OBLIGATOIRE :
Ta réponse doit être UNIQUEMENT un objet JSON valide, sans aucun texte avant ni après.
Commence directement par {{ et termine par }}. Aucun markdown, aucune explication hors JSON."""

    messages = [{"role": "system", "content": system_prompt}]
    for h in data.history[-8:]:
        if h.get("role") in ("user", "assistant") and h.get("content"):
            messages.append({"role": h["role"], "content": str(h["content"])})
    messages.append({"role": "user", "content": data.message})

    try:
        parsed = call_llm_json(messages, max_tokens=2048, temperature=0.2)

        action = parsed.get("action", "answer")

        # ── patch_transaction : toujours via confirmation frontend ────
        if action == "patch_transaction":
            parsed["confirm_endpoint"] = "/finance/transactions/" + str((parsed.get("data") or {}).get("id", 0))
            parsed["confirm_method"]   = "PATCH"

        # ── Exécuter directement si pas de confirmation requise ──────
        if not parsed.get("requires_confirmation", True) and action in ("add_transaction", "create_wallet"):
            d = parsed.get("data", {})
            if action == "add_transaction" and d:
                wallet = db.query(Wallet).filter(Wallet.id == d.get("wallet_id")).first()
                if wallet:
                    if d["type"] == "income":
                        wallet.balance += d["amount"]
                    else:
                        wallet.balance -= d["amount"]
                    tx = Transaction(
                        type=d["type"], amount=d["amount"],
                        category=d.get("category", "autre"),
                        description=d.get("description"),
                        wallet_id=d["wallet_id"], date=datetime.now(timezone.utc),
                    )
                    db.add(tx)
                    db.flush()  # get tx.id before commit

                    # Auto-match: provisional expense fulfilled → delete it
                    prov_id = d.get("provisional_id")
                    if prov_id:
                        prov = db.query(ProvisionalExpense).filter(ProvisionalExpense.id == prov_id).first()
                        if prov:
                            db.delete(prov)

                    # Auto-match: fixed charge paid → mark description
                    fc_id = d.get("fixed_charge_id")
                    if fc_id:
                        charge = db.query(FixedCharge).filter(FixedCharge.id == fc_id).first()
                        if charge:
                            tx.description = f"[Auto] {charge.name}"

                    db.commit()
            elif action == "create_wallet" and d:
                w = Wallet(name=d["name"], type=d.get("type", "autre"), balance=d.get("balance", 0))
                db.add(w)
                db.commit()

        # ── transfer → exécuter direct ou via frontend ───────────────
        if action == "transfer":
            parsed["confirm_endpoint"] = "/finance/transfer"
            if not parsed.get("requires_confirmation", True):
                d = parsed.get("data", {}) or {}
                src = db.query(Wallet).filter(Wallet.id == d.get("from_wallet_id")).first()
                dst = db.query(Wallet).filter(Wallet.id == d.get("to_wallet_id")).first()
                if src and dst and src.balance >= d.get("amount", 0):
                    src.balance -= d["amount"]
                    dst.balance += d["amount"]
                    desc = d.get("description") or f"Transfert → {dst.name}"
                    db.add(Transaction(type="expense", amount=d["amount"], category="transfert",
                        description=desc, wallet_id=src.id, date=datetime.now(timezone.utc)))
                    db.add(Transaction(type="income", amount=d["amount"], category="transfert",
                        description=desc, wallet_id=dst.id, date=datetime.now(timezone.utc)))
                    db.commit()

        # ── update_settings → appliquer directement ─────────────────
        if action == "update_settings":
            parsed["confirm_endpoint"] = "__update_settings__"

        # ── add_fixed_charge → hint endpoint ────────────────────────
        if action == "add_fixed_charge":
            parsed["confirm_endpoint"] = "/finance/fixed-charges"

        # ── update_provisional_expense ───────────────────────────────
        if action == "update_provisional_expense" and not parsed.get("requires_confirmation", True):
            d = parsed.get("data", {}) or {}
            p = db.query(ProvisionalExpense).filter(ProvisionalExpense.id == d.get("id")).first()
            if p:
                if d.get("description"): p.description = d["description"]
                if d.get("amount"):      p.amount = d["amount"]
                if d.get("category"):    p.category = d["category"]
                if d.get("month"):       p.month = d["month"]
                db.commit()
        if action == "update_provisional_expense":
            parsed["confirm_endpoint"] = "__update_provisional__"

        # ── delete_provisional_expense ───────────────────────────────
        if action == "delete_provisional_expense":
            parsed["confirm_endpoint"] = "__delete_provisional__"

        # ── add_provisional_expense → hint endpoint ──────────────────
        if action == "add_provisional_expense":
            parsed["confirm_endpoint"] = "/finance/provisional-expenses"
            # S'assurer que month est défini
            if parsed.get("data") and not parsed["data"].get("month"):
                from datetime import date
                today = date.today()
                if today.month == 12:
                    parsed["data"]["month"] = f"{today.year + 1}-01"
                else:
                    parsed["data"]["month"] = f"{today.year}-{today.month + 1:02d}"

        # ── Générer le PDF facture/devis ─────────────────────────────
        if action in ("generate_invoice", "generate_devis"):
            d = parsed.get("data", {})
            if d:
                pdf_b64, filename = _generate_pdf(d)
                parsed["pdf_base64"] = pdf_b64
                parsed["filename"] = filename

        # ── Générer un rapport PDF de données financières ────────────
        if action == "generate_report":
            d = parsed.get("data", {}) or {}
            pdf_b64, filename = _generate_report_pdf(d, db)
            parsed["pdf_base64"] = pdf_b64
            parsed["filename"] = filename
            parsed["requires_confirmation"] = False

        # ── Générer un Excel ─────────────────────────────────────────
        if action == "generate_excel":
            d = parsed.get("data", {}) or {}
            excel_b64, filename = _generate_excel(d, db)
            parsed["excel_base64"] = excel_b64
            parsed["filename"] = filename
            parsed["requires_confirmation"] = False

        # ── Auto-log ─────────────────────────────────────────────────
        try:
            from app.models.action_log import ActionLog
            _result = "pending_confirm" if parsed.get("requires_confirmation") else "auto"
            _log = ActionLog(
                prompt=data.message,
                ai_response=parsed.get("message", ""),
                action_type=parsed.get("action", "answer"),
                action_data=json.dumps(parsed.get("data")) if parsed.get("data") else None,
                result=_result,
            )
            db.add(_log)
            db.commit()
        except Exception:
            pass

        return parsed

    except json.JSONDecodeError:
        return {
            "action": "answer",
            "message": "Je n'ai pas pu interpréter la réponse. Reformule ta demande.",
            "requires_confirmation": False,
            "data": None,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur LLM : {str(e)}")


# ─── Action Logs ──────────────────────────────────────────────────────
class ActionLogCreate(BaseModel):
    prompt:      Optional[str] = None
    ai_response: Optional[str] = None
    action_type: Optional[str] = None
    action_data: Optional[str] = None
    result:      Optional[str] = None

@router.get("/action-logs/export/csv")
def export_action_logs_csv(db: Session = Depends(get_db)):
    import csv as _csv
    from app.models.action_log import ActionLog
    logs = db.query(ActionLog).order_by(ActionLog.timestamp.desc()).all()
    buf = io.StringIO()
    writer = _csv.writer(buf)
    writer.writerow(["id", "timestamp", "action_type", "result", "prompt", "ai_response", "action_data"])
    for l in logs:
        writer.writerow([
            l.id,
            l.timestamp.isoformat() if l.timestamp else "",
            l.action_type or "",
            l.result or "",
            (l.prompt or "").replace("\n", " "),
            (l.ai_response or "").replace("\n", " "),
            l.action_data or "",
        ])
    buf.seek(0)
    filename = f"historique_actions_{datetime.now().strftime('%Y-%m-%d')}.csv"
    return StreamingResponse(
        iter([buf.getvalue().encode("utf-8-sig")]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )

@router.get("/action-logs/export/pdf")
def export_action_logs_pdf(db: Session = Depends(get_db)):
    from app.models.action_log import ActionLog
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm

    logs = db.query(ActionLog).order_by(ActionLog.timestamp.desc()).limit(200).all()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    cell_s = ParagraphStyle("cell", parent=styles["Normal"], fontSize=7, leading=9)
    story  = []
    story.append(Paragraph("Historique des actions", ParagraphStyle("h1", parent=styles["Heading1"], fontSize=16, spaceAfter=8)))
    story.append(Paragraph(f"Exporté le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", styles["Normal"]))
    story.append(Spacer(1, 10))

    rows = [["Date", "Action", "Résultat", "Prompt", "Réponse IA"]]
    for l in logs:
        ts = l.timestamp.strftime("%d/%m %H:%M") if l.timestamp else ""
        rows.append([
            Paragraph(ts, cell_s),
            Paragraph(l.action_type or "-", cell_s),
            Paragraph(l.result or "-", cell_s),
            Paragraph((l.prompt or "")[:250], cell_s),
            Paragraph((l.ai_response or "")[:250], cell_s),
        ])
    tbl = Table(rows, colWidths=[22*mm, 25*mm, 18*mm, 62*mm, 62*mm])
    tbl.setStyle(TableStyle([
        ("BACKGROUND",   (0,0), (-1,0), colors.HexColor("#1e1e2e")),
        ("TEXTCOLOR",    (0,0), (-1,0), colors.white),
        ("FONTNAME",     (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",     (0,0), (-1,0), 8),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white, colors.HexColor("#f5f5f8")]),
        ("GRID",         (0,0), (-1,-1), 0.25, colors.HexColor("#cccccc")),
        ("VALIGN",       (0,0), (-1,-1), "TOP"),
        ("TOPPADDING",   (0,0), (-1,-1), 3),
        ("BOTTOMPADDING",(0,0), (-1,-1), 3),
    ]))
    story.append(tbl)
    doc.build(story)
    buf.seek(0)
    filename = f"historique_actions_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})

@router.get("/action-logs")
def get_action_logs(limit: int = 200, db: Session = Depends(get_db)):
    from app.models.action_log import ActionLog
    logs = db.query(ActionLog).order_by(ActionLog.timestamp.desc()).limit(limit).all()
    return [
        {
            "id":          l.id,
            "timestamp":   l.timestamp.isoformat() if l.timestamp else None,
            "prompt":      l.prompt,
            "ai_response": l.ai_response,
            "action_type": l.action_type,
            "action_data": l.action_data,
            "result":      l.result,
        }
        for l in logs
    ]

@router.post("/action-logs")
def create_action_log(data: ActionLogCreate, db: Session = Depends(get_db)):
    from app.models.action_log import ActionLog
    log = ActionLog(
        prompt=data.prompt,
        ai_response=data.ai_response,
        action_type=data.action_type,
        action_data=data.action_data,
        result=data.result,
    )
    db.add(log)
    db.commit()
    db.refresh(log)
    return {"id": log.id}


# ─── Generate Invoice PDF ─────────────────────────────────────────
class InvoiceRequest(BaseModel):
    title: str
    client: str
    items: List[dict]
    notes: str = ""
    type: str = "facture"   # "facture" ou "devis"


@router.post("/generate-invoice")
def generate_invoice(data: InvoiceRequest):
    """Génère un PDF de facture ou devis et retourne le base64."""
    payload = data.model_dump()
    pdf_b64, filename = _generate_pdf(payload)
    return {"pdf_base64": pdf_b64, "filename": filename}


def _generate_pdf(data: dict) -> tuple[str, str]:
    """Génère un PDF professionnel et retourne (base64, filename)."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_RIGHT, TA_CENTER

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4,
                                topMargin=2*cm, bottomMargin=2*cm,
                                leftMargin=2*cm, rightMargin=2*cm)

        styles = getSampleStyleSheet()
        dark = colors.HexColor('#1a1a2e')
        accent = colors.HexColor('#6ee7b7')
        light_grey = colors.HexColor('#f8f9fa')
        mid_grey = colors.HexColor('#6c757d')

        doc_type = data.get("type", "facture").upper()
        title = data.get("title", "Document")
        client = data.get("client", "")
        items = data.get("items", [])
        notes = data.get("notes", "")
        doc_num = f"{doc_type[:3]}-{datetime.now().strftime('%Y%m%d%H%M')}"
        date_str = datetime.now().strftime('%d/%m/%Y')

        story = []

        # En-tête
        header_data = [
            [Paragraph(f"<font size='22' color='#{dark.hexval()[2:]}'><b>{doc_type}</b></font>",
                       styles['Normal']),
             Paragraph(f"<font size='10' color='#{mid_grey.hexval()[2:]}'>N° {doc_num}<br/>Date : {date_str}</font>",
                       ParagraphStyle('right', parent=styles['Normal'], alignment=TA_RIGHT))
             ]
        ]
        header_table = Table(header_data, colWidths=[10*cm, 7*cm])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ]))
        story.append(header_table)
        story.append(Spacer(1, 0.5*cm))

        # Titre du document
        story.append(Paragraph(
            f"<font size='14' color='#1a1a2e'><b>{title}</b></font>",
            styles['Normal']
        ))
        story.append(Spacer(1, 0.3*cm))

        # Client
        if client:
            story.append(Paragraph(
                f"<font size='11' color='#{mid_grey.hexval()[2:]}'>Client : </font>"
                f"<font size='11' color='#1a1a2e'><b>{client}</b></font>",
                styles['Normal']
            ))
        story.append(Spacer(1, 0.6*cm))

        # Tableau des articles
        table_data = [['Article', 'Qté', 'Prix unitaire', 'Total']]
        total = 0
        for item in items:
            name = str(item.get('name', ''))
            qty = float(item.get('qty', 1))
            unit_price = int(item.get('unit_price', 0))
            line_total = int(qty * unit_price)
            total += line_total
            table_data.append([
                name,
                f"{qty:g}",
                f"{unit_price:,} Ar".replace(',', ' '),
                f"{line_total:,} Ar".replace(',', ' '),
            ])

        # Ligne total
        table_data.append(['', '', 'TOTAL', f"{total:,} Ar".replace(',', ' ')])

        n_items = len(table_data)
        col_w = [8.5*cm, 2*cm, 3.5*cm, 3*cm]
        t = Table(table_data, colWidths=col_w, repeatRows=1)
        t.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), dark),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (1, 0), (-1, 0), 'RIGHT'),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            # Corps
            ('FONTSIZE', (0, 1), (-1, -2), 10),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ('TOPPADDING', (0, 1), (-1, -2), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -2), 6),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [light_grey, colors.white]),
            # Ligne total
            ('BACKGROUND', (0, n_items-1), (-1, n_items-1), colors.HexColor('#e8faf4')),
            ('FONTNAME', (0, n_items-1), (-1, n_items-1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, n_items-1), (-1, n_items-1), 11),
            ('TEXTCOLOR', (2, n_items-1), (-1, n_items-1), colors.HexColor('#059669')),
            ('TOPPADDING', (0, n_items-1), (-1, n_items-1), 8),
            ('BOTTOMPADDING', (0, n_items-1), (-1, n_items-1), 8),
            # Bordures
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
            ('LINEBELOW', (0, 0), (-1, 0), 1.5, accent),
        ]))
        story.append(t)

        if notes:
            story.append(Spacer(1, 0.6*cm))
            story.append(Paragraph(
                f"<font size='10' color='#{mid_grey.hexval()[2:]}'>Notes : {notes}</font>",
                styles['Normal']
            ))

        story.append(Spacer(1, 1*cm))
        story.append(Paragraph(
            f"<font size='9' color='#{mid_grey.hexval()[2:]}'>Document généré le {date_str} — MyLife Financial</font>",
            ParagraphStyle('footer', parent=styles['Normal'], alignment=TA_CENTER)
        ))

        doc.build(story)
        pdf_bytes = buf.getvalue()
        pdf_b64 = base64.b64encode(pdf_bytes).decode('utf-8')
        filename = f"{doc_num}.pdf"
        return pdf_b64, filename

    except ImportError:
        raise HTTPException(status_code=503, detail="reportlab non installé — pip install reportlab")


# ─── Statistiques mensuelles (graphe tendances) ───────────────────
@router.get("/monthly-stats")
def monthly_stats(months: int = 6, db: Session = Depends(get_db)):
    """Revenus et dépenses totaux pour les N derniers mois."""
    import calendar
    now = datetime.now(timezone.utc)
    month_names = ['Jan','Fév','Mar','Avr','Mai','Juin','Juil','Août','Sep','Oct','Nov','Déc']
    result = []
    for i in range(months - 1, -1, -1):
        month = now.month - i
        year  = now.year
        while month <= 0:
            month += 12
            year  -= 1
        start = datetime(year, month, 1, tzinfo=timezone.utc)
        last_day = calendar.monthrange(year, month)[1]
        end   = datetime(year, month, last_day, 23, 59, 59, tzinfo=timezone.utc)
        income = db.query(func.sum(Transaction.amount)).filter(
            Transaction.type == "income",
            Transaction.date >= start, Transaction.date <= end,
        ).scalar() or 0
        expenses = db.query(func.sum(Transaction.amount)).filter(
            Transaction.type == "expense",
            Transaction.date >= start, Transaction.date <= end,
        ).scalar() or 0
        result.append({"month": month_names[month - 1], "income": income, "expenses": expenses})
    return result


# ─── Smart tip LLM (expert financier, cache 30 min) ──────────────
@router.get("/smart-tip")
def smart_tip(db: Session = Depends(get_db)):
    """Retourne un conseil financier expert généré par LLM, mis en cache 30 min."""
    global _tip_cache
    if _tip_cache["tip"] and time.time() - _tip_cache["ts"] < 1800:
        return _tip_cache

    now = datetime.now(timezone.utc)
    start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    income = db.query(func.sum(Transaction.amount)).filter(
        Transaction.type == "income", Transaction.date >= start
    ).scalar() or 0
    expenses = db.query(func.sum(Transaction.amount)).filter(
        Transaction.type == "expense", Transaction.date >= start
    ).scalar() or 0
    wallets = db.query(Wallet).all()
    total = sum(w.balance for w in wallets)

    cat_map: dict = {}
    txs = db.query(Transaction).filter(Transaction.date >= start).all()
    for tx in txs:
        if tx.type == "expense":
            cat_map[tx.category] = cat_map.get(tx.category, 0) + tx.amount

    top_cat = max(cat_map, key=cat_map.get) if cat_map else None
    ratio = round(expenses / income * 100) if income > 0 else None

    prompt = f"""Expert financier personnel. Analyse ces données du mois en cours et donne UN conseil court (2 phrases max).
Style : direct, chiffré, comme un vrai conseiller. Pas de banalités.

Données :
- Revenus ce mois : {income:,} Ar
- Dépenses ce mois : {expenses:,} Ar
- Ratio dépenses/revenus : {ratio}%
- Solde total : {total:,} Ar
- Catégorie dépenses principale : {top_cat} ({cat_map.get(top_cat, 0):,} Ar)

Réponds en JSON uniquement : {{"tip": "...", "type": "warning|info|good"}}
warning si ratio > 80%, good si ratio < 50%, info sinon."""

    try:
        parsed = call_llm_json([{"role": "user", "content": prompt}], max_tokens=120, temperature=0.6)
        _tip_cache = {"tip": parsed.get("tip", ""), "type": parsed.get("type", "info"), "ts": time.time()}
        return _tip_cache
    except Exception:
        return {"tip": f"Taux d'effort : {ratio}% de tes revenus en dépenses ce mois.", "type": "warning" if (ratio or 0) > 80 else "info"}


# ─── Export Excel ─────────────────────────────────────────────────
def _generate_excel(data: dict, db: Session) -> tuple[str, str]:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    data_type = data.get("type", "transactions")
    wb = Workbook()

    dark_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)

    def style_header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = dark_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

    def auto_width(ws):
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, 40)

    if data_type in ("transactions", "full", "monthly_report"):
        ws = wb.active
        ws.title = "Transactions"
        headers = ["Date", "Type", "Montant (Ar)", "Catégorie", "Description", "Portefeuille"]
        style_header(ws, headers)
        txs = db.query(Transaction).order_by(Transaction.date.desc()).limit(500).all()
        for r, tx in enumerate(txs, 2):
            ws.cell(r, 1, tx.date.strftime("%d/%m/%Y"))
            ws.cell(r, 2, "Revenu" if tx.type == "income" else "Dépense")
            ws.cell(r, 3, tx.amount)
            ws.cell(r, 4, tx.category)
            ws.cell(r, 5, tx.description or "")
            ws.cell(r, 6, tx.wallet.name if tx.wallet else "")
        auto_width(ws)

    if data_type in ("wallets", "full"):
        ws2 = wb.create_sheet("Portefeuilles")
        style_header(ws2, ["Nom", "Type", "Solde (Ar)"])
        for r, w in enumerate(db.query(Wallet).all(), 2):
            ws2.cell(r, 1, w.name); ws2.cell(r, 2, w.type); ws2.cell(r, 3, w.balance)
        auto_width(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"finances_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return base64.b64encode(buf.getvalue()).decode(), filename


# ─── Rapport PDF données financières ─────────────────────────────
def _generate_report_pdf(data: dict, db: Session) -> tuple[str, str]:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm,
                            leftMargin=2*cm, rightMargin=2*cm)
    styles = getSampleStyleSheet()
    dark = colors.HexColor('#1a1a2e')
    story = []

    title = data.get("title", "Rapport financier")
    months = data.get("months", 1)
    story.append(Paragraph(f"<b>{title}</b>", styles['Title']))
    story.append(Paragraph(datetime.now().strftime("%d/%m/%Y %H:%M"), styles['Normal']))
    story.append(Spacer(1, 0.5*cm))

    # Wallets summary
    wallets = db.query(Wallet).all()
    w_data = [["Portefeuille", "Type", "Solde"]]
    for w in wallets:
        w_data.append([w.name, w.type, f"{w.balance:,} Ar".replace(',', ' ')])
    t = Table(w_data, colWidths=[7*cm, 4*cm, 5*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), dark), ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'), ('FONTSIZE', (0,0), (-1,-1), 10),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#f8f9fa'), colors.white]),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#dee2e6')),
        ('TOPPADDING', (0,0), (-1,-1), 6), ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.5*cm))

    # Transactions
    txs = db.query(Transaction).order_by(Transaction.date.desc()).limit(100).all()
    story.append(Paragraph("<b>Dernières transactions</b>", styles['Heading2']))
    story.append(Spacer(1, 0.2*cm))
    tx_data = [["Date", "Type", "Montant", "Catégorie", "Portefeuille"]]
    for tx in txs:
        tx_data.append([
            tx.date.strftime("%d/%m/%Y"),
            "+" if tx.type == "income" else "-",
            f"{tx.amount:,} Ar".replace(',', ' '),
            tx.category,
            tx.wallet.name if tx.wallet else "",
        ])
    t2 = Table(tx_data, colWidths=[2.5*cm, 1.5*cm, 3.5*cm, 3.5*cm, 4*cm])
    t2.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), dark), ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'), ('FONTSIZE', (0,0), (-1,-1), 8),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#f8f9fa'), colors.white]),
        ('GRID', (0,0), (-1,-1), 0.3, colors.HexColor('#dee2e6')),
        ('TOPPADDING', (0,0), (-1,-1), 4), ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(t2)
    doc.build(story)
    buf.seek(0)
    filename = f"rapport_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return base64.b64encode(buf.getvalue()).decode(), filename


# ─── Transcription audio → texte (Whisper) ───────────────────────
@router.post("/transcribe")
async def transcribe_audio(audio: UploadFile = File(...)):
    """Reçoit un fichier audio (webm/wav/mp4) et retourne le texte transcrit via Whisper."""
    keys = _openai_keys()
    if not keys:
        raise HTTPException(status_code=503, detail="Pas de clé OpenAI pour Whisper")

    audio_bytes = await audio.read()
    filename = audio.filename or "audio.webm"
    content_type = audio.content_type or "audio/webm"

    try:
        resp = httpx.post(
            "https://api.openai.com/v1/audio/transcriptions",
            headers={"Authorization": f"Bearer {keys[0]}"},
            files={"file": (filename, audio_bytes, content_type)},
            data={"model": "whisper-1", "language": "fr"},
            timeout=30.0,
        )
        resp.raise_for_status()
        return {"text": resp.json().get("text", "")}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur Whisper : {e}")


# ─── Voice / LLM ─────────────────────────────────────────────────
class VoiceRequest(BaseModel):
    text: str


@router.post("/transcribe")
async def transcribe_audio(file: UploadFile = File(...)):
    """Transcrit un fichier audio via Whisper (OpenAI). Retourne {"text": "..."}"""
    keys = []
    for k in ("OPENAI_API_KEY_1", "OPENAI_API_KEY_2", "OPENAI_API_KEY"):
        v = os.getenv(k, "").strip()
        if v and v not in keys:
            keys.append(v)
    if not keys:
        raise HTTPException(status_code=503, detail="Pas de clé OpenAI configurée")

    audio_bytes = await file.read()
    filename = file.filename or "audio.m4a"

    for key in keys:
        try:
            async with httpx.AsyncClient(timeout=30) as client:
                r = await client.post(
                    "https://api.openai.com/v1/audio/transcriptions",
                    headers={"Authorization": f"Bearer {key}"},
                    data={"model": "whisper-1", "language": "fr"},
                    files={"file": (filename, audio_bytes, file.content_type or "audio/m4a")},
                )
            if r.status_code == 200:
                return {"text": r.json().get("text", "").strip()}
        except Exception:
            continue

    raise HTTPException(status_code=500, detail="Transcription Whisper échouée")


@router.post("/voice")
def parse_voice(data: VoiceRequest, db: Session = Depends(get_db)):
    """Reçoit du texte (STT) et utilise un LLM pour extraire les infos de transaction."""
    wallets = db.query(Wallet).all()
    wallet_list = "\n".join(
        f'- id={w.id}, nom="{w.name}", type={w.type}, solde={w.balance} Ar'
        for w in wallets
    )

    prompt = f"""Tu es un assistant financier pour un utilisateur malgache.
Analyse ce texte vocal et extrais les informations de transaction.

Portefeuilles disponibles :
{wallet_list}

RÈGLES DE WALLET (priorité absolue) :
- Si aucun wallet mentionné → utiliser "Argent liquide" (wallet par défaut)
- ALIMENTATION / NOURRITURE : repas, courses, épicerie, riz, restaurant, marché → wallet = "Argent liquide" TOUJOURS
- "crédit Yas", "recharge Yas", "forfait Yas" → wallet = "Mvola" par défaut, SAUF si l'utilisateur précise l'origine (ex: "en liquide", "cash" → Argent liquide)
- "connexion Orange", "internet Orange", "forfait Orange" → wallet = "Orange Money"
- "retrait Mvola", "retrait Orange Money" → ce sont des transferts (transfer), pas des expenses

Texte vocal : "{data.text}"

Réponds UNIQUEMENT avec un objet JSON valide (sans markdown) avec ces champs :
- type : "income" ou "expense"
- amount : nombre entier en Ariary (extrait le montant du texte, null si absent)
- category : une de ces catégories : alimentation, transport, salaire, transfert, loyer, santé, loisirs, mvola, orange_money, internet, autre
- wallet_id : l'id du portefeuille le plus adapté (parmi les ids disponibles)
- description : courte description de la transaction
- confidence : "high", "medium" ou "low"

Exemples d'interprétation :
- "j'ai reçu mon salaire sur mon compte" → type=income, category=salaire, wallet=compte bancaire
- "j'ai fait un achat depuis mvola de 10000" → type=expense, category=transfert, wallet=Mvola, amount=10000
- "j'ai dépensé 5000 en nourriture" → type=expense, category=alimentation, wallet=Argent liquide
- "j'ai acheté un crédit Yas de 3000" → type=expense, category=internet, wallet=Mvola
- "j'ai acheté un crédit Yas de 3000 en liquide" → type=expense, category=internet, wallet=Argent liquide
- "j'ai payé ma connexion Orange" → type=expense, category=internet, wallet=Orange Money"""

    try:
        parsed = call_llm_json(
            [{"role": "user", "content": prompt}],
            max_tokens=256, temperature=0.1,
        )
        w = next((w for w in wallets if w.id == parsed.get("wallet_id")), None)
        parsed["wallet_name"] = w.name if w else ""
        return parsed
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur LLM : {str(e)}")
